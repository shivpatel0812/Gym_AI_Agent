#!/usr/bin/env python3
"""
Python script to replicate the VB.NET Excel button functionality.
This script creates an Excel file with a button-like cell and VBA macro.
"""

import sys
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("openpyxl is not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_excel_with_button():
    """Creates an Excel file with a button-like cell"""
    try:
        # Create a new workbook
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"
        
        # Add instruction text
        sheet["A1"].value = "Click the button below:"
        sheet["A1"].font = Font(bold=True, size=12)
        
        # Set cell A2 (where the message will appear when button is clicked)
        cell_a2 = sheet["A2"]
        cell_a2.value = ""  # Will be set by button click
        cell_a2.font = Font(bold=True, size=12, color="0066CC")
        cell_a2.alignment = Alignment(horizontal="left", vertical="center")
        
        # Create a button-like cell (B1) with blue background
        button_cell = sheet["B1"]
        button_cell.value = "Click Me!"
        button_cell.font = Font(bold=True, size=14, color="FFFFFF")
        button_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        button_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add border to make it look more like a button
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        button_cell.border = thin_border
        
        # Adjust row and column sizes
        sheet.row_dimensions[1].height = 30
        sheet.column_dimensions['B'].width = 15
        
        # Save the workbook
        output_file = "button_output.xlsx"
        wb.save(output_file)
        
        # Create VBA macro file
        create_vba_macro()
        
        abs_path = os.path.abspath(output_file)
        
        print("✓ Successfully created Excel file with button!")
        print(f"✓ File saved as: {abs_path}")
        print(f"\n📝 To make the button work:")
        print(f"   1. Open '{output_file}' in Excel")
        print(f"   2. Press Alt+F11 (or Cmd+Option+F11 on Mac) to open VBA Editor")
        print(f"   3. Right-click 'Sheet1' in the left panel")
        print(f"   4. Select 'View Code'")
        print(f"   5. Paste the code from 'Sheet1.vba' file")
        print(f"   6. Close VBA Editor")
        print(f"   7. Right-click the blue 'Click Me!' button in cell B1")
        print(f"   8. Select 'Assign Macro' > 'Button1_Click'")
        print(f"   9. Click OK")
        print(f"\n   Now clicking the button will set cell A2!")
        
        return abs_path
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

def create_vba_macro():
    """Creates a VBA macro file that can be imported into Excel"""
    vba_code = '''Private Sub Button1_Click()
    Range("A2").Value = "Hi Ananya Manikandan. Amazing coding job!"
End Sub
'''
    
    vba_file = "Sheet1.vba"
    with open(vba_file, 'w') as f:
        f.write(vba_code)
    
    print(f"✓ VBA code saved to: {vba_file}")

if __name__ == "__main__":
    create_excel_with_button()
